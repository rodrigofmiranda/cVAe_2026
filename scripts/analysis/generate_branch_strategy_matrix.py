#!/usr/bin/env python3
"""Generate a teaching-oriented Excel workbook for branch/workline strategy review.

The workbook is intended for beginner students onboarding into the VLC digital
twin project.  It answers:

1. Which branches / work lines exist?
2. What strategic hypothesis each branch tested?
3. How far each architecture family was explored?
4. Which families are anchors, open questions, or closed negatives?
"""

from __future__ import annotations

import subprocess
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_XLSX = REPO_ROOT / "docs" / "active" / "DIGITAL_TWIN_BRANCH_STRATEGY_MATRIX.xlsx"


@dataclass(frozen=True)
class BranchLine:
    branch_name: str
    metadata_ref: str
    worktree_path: str
    strategic_role: str
    strategic_hypothesis: str
    architecture_core: str
    generator_family: str
    tests_realizados: str
    best_result: str
    latest_result: str
    extent_code: str
    verdict_flag: str
    digital_twin_lesson: str
    docs_entrypoint: str
    notes: str


@dataclass(frozen=True)
class ArchitectureLine:
    architecture_family: str
    name_breakdown_and_data_use: str
    main_branches: str
    objective_for_digital_twin: str
    interventions_tested: str
    farthest_stage_code: str
    best_verified_result: str
    current_status: str
    what_we_learned: str
    reopen_rule: str


@dataclass(frozen=True)
class ExperimentRow:
    run_id: str
    branch_line: str
    architecture_family: str
    preset_or_probe: str
    purpose: str
    pass_result: str
    gates: str
    verdict: str
    note: str


BRANCH_LINES = [
    BranchLine(
        branch_name="exp/refactor_architecture",
        metadata_ref="exp/refactor_architecture",
        worktree_path="historical branch only",
        strategic_role="Engineering base that modernized the pipeline before the current architecture races.",
        strategic_hypothesis="Separate platform modernization from scientific architecture bets so later protocol comparisons are trustworthy.",
        architecture_core="platform base for concat/residual/sequence families",
        generator_family="platform / protocol infrastructure branch",
        tests_realizados="protocol maturation; train_once_eval_all; grid diagnostics; stronger plots and reports",
        best_result="Not a direct protocol winner branch; enabled later scientific anchors.",
        latest_result="Retained as the clean engineering base behind later release/research lines.",
        extent_code="E5",
        verdict_flag="SUPPORT_BASELINE",
        digital_twin_lesson="A digital twin project needs a stable platform branch, not only model branches.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/archive/branching/BRANCH_STRATEGY.md",
        notes="Pedagogically important because it explains why release and research were separated later.",
    ),
    BranchLine(
        branch_name="release/cvae-online",
        metadata_ref="release/cvae-online",
        worktree_path="historical branch only",
        strategic_role="Functional online cVAE release branch kept separate from unresolved research routes.",
        strategic_hypothesis="A deployable/teachable online baseline should not depend on whichever research hypothesis is currently open.",
        architecture_core="functional online cVAE baseline",
        generator_family="release / deployment branch",
        tests_realizados="release-base validation; recommended clone path for functional online use",
        best_result="Recommended functional base, not the main current scientific frontier.",
        latest_result="Stable release-oriented branch kept for operational clarity.",
        extent_code="E5",
        verdict_flag="SUPPORT_BASELINE",
        digital_twin_lesson="Students should learn to separate a working online baseline from the branch chasing the final digital twin.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/README.md",
        notes="Not a research winner branch; it exists so the project remains usable while research continues.",
    ),
    BranchLine(
        branch_name="feat/channel-residual-architecture",
        metadata_ref="feat/channel-residual-architecture",
        worktree_path="historical branch only",
        strategic_role="Residual-architecture branch that preserved the channel_residual strategy without replacing the older concat path.",
        strategic_hypothesis="A residual architecture could model the channel more faithfully than the original concat family while keeping the baseline path intact.",
        architecture_core="channel_residual",
        generator_family="residual cVAE family",
        tests_realizados="residual-architecture branch establishment; baseline coexistence with concat; historical residual strategy work",
        best_result="Serious architectural line preserved for provenance; not the strongest current anchor.",
        latest_result="Kept as the recorded residual-architecture strategy, not the active scientific frontier.",
        extent_code="E4",
        verdict_flag="IMPLEMENTED_NOT_PROMOTED",
        digital_twin_lesson="A branch can represent a real architecture family even if later branches move the frontier elsewhere.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/archive/branching/BRANCH_STRATEGY.md",
        notes="Good teaching example of an architectural branch that remains important historically even when it is not the current focus.",
    ),
    BranchLine(
        branch_name="feat/seq-bigru-residual-cvae",
        metadata_ref="feat/seq-bigru-residual-cvae",
        worktree_path="historical parent research branch",
        strategic_role="Unified research branch that carried the seq_bigru_residual and delta_residual research line before the later specialization branches split off.",
        strategic_hypothesis="Sequence-aware residual modeling is the main route toward the final digital twin and deserves its own research lane apart from release branches.",
        architecture_core="seq_bigru_residual and delta_residual umbrella research lane",
        generator_family="research umbrella for Gaussian/MDN descendants",
        tests_realizados="daily seq research lineage; initial seq experiments; branch-level status consolidation before later branch specialization",
        best_result="Parent line behind later Gaussian 10/12 and MDN 9/12-10/12 descendants.",
        latest_result="Acts as the root research lineage reference rather than the newest specialized frontier branch.",
        extent_code="E5",
        verdict_flag="ROOT_RESEARCH_LINE",
        digital_twin_lesson="This is the branch type students should recognize as a parent P&D lane: later branches split from it when the scientific questions become specific.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/README.md",
        notes="The current worktree name itself comes from this branch lineage.",
    ),
    BranchLine(
        branch_name="feat/delta-residual-adv",
        metadata_ref="feat/delta-residual-adv",
        worktree_path="historical branch only",
        strategic_role="Archived adversarial branch kept for traceability rather than active use.",
        strategic_hypothesis="Adversarial pressure might improve realism of the residual/noise distribution.",
        architecture_core="delta_residual adversarial lane",
        generator_family="adversarial / GAN-adjacent residual branch",
        tests_realizados="historical adversarial branch work retained only for traceability; not part of the active current-cycle search",
        best_result="Archived only; no current-cycle benchmark role.",
        latest_result="Kept only as traceability for an idea family that is not active now.",
        extent_code="E2",
        verdict_flag="ARCHIVED_TRACEABILITY",
        digital_twin_lesson="Students should learn that archived traceability branches are different from active open routes.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/README.md",
        notes="Relevant whenever beginners ask whether GAN/adversarial routes were ever considered.",
    ),
    BranchLine(
        branch_name="feat/seq-imdd-graybox-mdn",
        metadata_ref="feat/seq-imdd-graybox-mdn",
        worktree_path="/workspace/2026/feat_seq_bigru_residual_cvae",
        strategic_role="Implement and validate the seq_imdd_graybox + MDN route on top of the gray-box physical scaffold.",
        strategic_hypothesis="A physics-informed gray-box prior plus MDN output could copy the noise shape better than the gray-box Gaussian decoder.",
        architecture_core="seq_imdd_graybox",
        generator_family="MDN over gray-box residual decoder",
        tests_realizados="gray-box Gaussian anchors inherited; seq_imdd_graybox_mdn_smoke; seq_imdd_graybox_mdn_guided_quick",
        best_result="Gray-box Gaussian anchor 6/12 (exp_20260327_172148); gray-box+MDN best 5/12 (exp_20260328_023302)",
        latest_result="Implemented and tested end-to-end; not promoted above gray-box Gaussian or seq MDN anchor.",
        extent_code="E4",
        verdict_flag="IMPLEMENTED_NOT_PROMOTED",
        digital_twin_lesson="Adding MDN solved the implementation gap for gray-box, but not the global residual-shape gap.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/active/ROUTES_AND_RESULTS.md",
        notes="Useful as a physics-scaffold branch, but still below the strongest seq_bigru_residual MDN line.",
    ),
    BranchLine(
        branch_name="feat/imdd-graybox-channel",
        metadata_ref="feat/imdd-graybox-channel",
        worktree_path="historical branch (no dedicated current worktree)",
        strategic_role="Gray-box Gaussian lane with guided capacity/context sweeps.",
        strategic_hypothesis="A structured IMDD gray-box decoder might recover the channel with less purely learned complexity.",
        architecture_core="seq_imdd_graybox",
        generator_family="Gaussian gray-box decoder",
        tests_realizados="smoke; capacity quick; guided large",
        best_result="6/12 on exp_20260327_172148",
        latest_result="5/12 on exp_20260327_183153 after a larger guided grid.",
        extent_code="E4",
        verdict_flag="VIABLE_BELOW_ANCHOR",
        digital_twin_lesson="Physics structure helps, but Gaussian gray-box still misses the global shape of the noise.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/active/ROUTES_AND_RESULTS.md",
        notes="Important for students because it shows a physics-guided branch that is viable but not sufficient.",
    ),
    BranchLine(
        branch_name="feat/seq-bigru-residual-mdn-route",
        metadata_ref="feat/seq-bigru-residual-mdn-route",
        worktree_path="/workspace/2026/feat_seq_bigru_residual_mdn_route",
        strategic_role="Dedicated reproducibility and rerun lane for the historical seq_bigru_residual + MDN family.",
        strategic_hypothesis="If the old 9/12 MDN result is real and robust, it should reproduce cleanly under the current stack.",
        architecture_core="seq_bigru_residual",
        generator_family="MDN",
        tests_realizados="seq_mdn_v2_overnight_5090safe_quick rerun; exact-candidate isolation rerun",
        best_result="Historical benchmark 9/12 (exp_20260327_161311)",
        latest_result="4/12 (exp_20260328_041729) and 2/12 isolation (exp_20260328_181213)",
        extent_code="E4",
        verdict_flag="REPRO_GAP",
        digital_twin_lesson="Good-looking anchors are not enough; runtime stack, host provenance, and RNG discipline matter for a digital twin.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_mdn_route/docs/active/ROUTES_AND_RESULTS.md",
        notes="Main branch for teaching reproducibility drift and why we compare environments, not only hyperparameters.",
    ),
    BranchLine(
        branch_name="feat/mdn-g5-recovery",
        metadata_ref="origin/feat/mdn-g5-recovery",
        worktree_path="remote/main line; no synced dedicated local worktree right now",
        strategic_role="Main high-performance MDN branch for attacking the remaining G5 failures while preserving the seq MDN anchor.",
        strategic_hypothesis="Coverage/tail shaping and decoder-side conditioning can push the best seq_bigru_residual + MDN line beyond the current 10/12 ceiling or explain the ceiling.",
        architecture_core="seq_bigru_residual",
        generator_family="MDN + coverage/tail + decoder conditioning probes",
        tests_realizados="S27 lambda_coverage sweep; S28 lambda_kurt; S29 mdn_components; S30 cond_embed; S31 tuned cond_embed; S32 large decoder embed sweep (running per latest docs)",
        best_result="10/12 on exp_20260328_153611 (S27 coverage anchor)",
        latest_result="Remote docs updated on 2026-03-30 with S30 and S31 conclusions; S32 documented as the next broad embed sweep.",
        extent_code="E5",
        verdict_flag="ANCHOR_OPEN",
        digital_twin_lesson="This is the strongest current learned digital twin line: it wins on protocol, but still leaves a global residual-shape gap.",
        docs_entrypoint="origin/feat/mdn-g5-recovery:docs/active/WORKING_STATE.md",
        notes="Latest remote commit was authored by IA VLC CT102 / Claude-assisted flow on 2026-03-30T16:56:37Z.",
    ),
    BranchLine(
        branch_name="feat/mdn-g5-recovery-run",
        metadata_ref="feat/mdn-g5-recovery-run",
        worktree_path="/workspace/2026/feat_mdn_g5_recovery",
        strategic_role="Local snapshot/handoff lane for the regime-conditioning experiments.",
        strategic_hypothesis="Keep a stable local state for S30-era docs and protocol reading while the main recovery branch continues separately.",
        architecture_core="seq_bigru_residual",
        generator_family="MDN + cond_embed snapshot",
        tests_realizados="S30 regime-conditioning result recording; local handoff docs",
        best_result="5/12 on exp_20260328_233844",
        latest_result="Snapshot branch; superseded scientifically by the newer remote feat/mdn-g5-recovery branch.",
        extent_code="E3",
        verdict_flag="SNAPSHOT_SUPERSEDED",
        digital_twin_lesson="Useful for pedagogy: branches can be snapshots for handoff, not only active research frontiers.",
        docs_entrypoint="/workspace/2026/feat_mdn_g5_recovery/docs/active/WORKING_STATE.md",
        notes="Tracks origin/feat/mdn-g5-recovery-run, not the newer origin/feat/mdn-g5-recovery branch.",
    ),
    BranchLine(
        branch_name="feat/conditional-flow-decoder",
        metadata_ref="feat/conditional-flow-decoder",
        worktree_path="historical branch only",
        strategic_role="First historical conditional density decoder alternative beyond Gaussian/MDN.",
        strategic_hypothesis="A narrow per-axis conditional flow decoder could represent non-Gaussian residual shape more faithfully.",
        architecture_core="seq_bigru_residual / historical decoder experiment",
        generator_family="sinh-arcsinh flow",
        tests_realizados="historical protocol runs exp_20260326_034522 and exp_20260326_035723",
        best_result="0/12",
        latest_result="0/12; branch retained only as negative evidence.",
        extent_code="E4",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="Richer density family on paper is not enough; the chosen flow formulation must match the physics and the residual geometry.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/reference/CONDITIONAL_DENSITY_DECODER_GUIDE.md",
        notes="This is the old narrow per-axis flow line that should not be reopened blindly.",
    ),
    BranchLine(
        branch_name="feat/sample-aware-mmd",
        metadata_ref="feat/sample-aware-mmd",
        worktree_path="historical branch only",
        strategic_role="Sample-aware distribution matching line in the training objective.",
        strategic_hypothesis="Pushing MMD on sampled residuals could improve global distribution shape beyond mean-residual losses.",
        architecture_core="seq_bigru_residual / MDN-adjacent loss branch",
        generator_family="MDN or Gaussian with sample-aware MMD objective",
        tests_realizados="historical sample-aware MMD branch experiments",
        best_result="Negative historical family result",
        latest_result="Closed as negative.",
        extent_code="E3",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="A better training loss does not automatically fix a mismatched generative family.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/active/ROUTES_AND_RESULTS.md",
        notes="GPU and local variants below are operational variants of the same negative family.",
    ),
    BranchLine(
        branch_name="feat/sample-aware-mmd-gpu",
        metadata_ref="feat/sample-aware-mmd-gpu",
        worktree_path="historical branch only",
        strategic_role="GPU execution variant of the sample-aware MMD line.",
        strategic_hypothesis="Operational changes on GPU might rescue or clarify the MMD family.",
        architecture_core="seq_bigru_residual / MDN-adjacent loss branch",
        generator_family="sample-aware MMD variant",
        tests_realizados="historical GPU variant of sample-aware MMD",
        best_result="Negative family result",
        latest_result="Closed with the rest of the MMD line.",
        extent_code="E3",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="Operational variants are useful for debugging, but they do not replace a successful scientific formulation.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/active/ROUTES_AND_RESULTS.md",
        notes="Keep for provenance, not as an active route.",
    ),
    BranchLine(
        branch_name="feat/sample-aware-mmd-gpu-local",
        metadata_ref="feat/sample-aware-mmd-gpu-local",
        worktree_path="historical branch only",
        strategic_role="Local runtime variant of the sample-aware MMD family.",
        strategic_hypothesis="Local hardware/runtime could alter the behavior of the MMD route enough to matter.",
        architecture_core="seq_bigru_residual / MDN-adjacent loss branch",
        generator_family="sample-aware MMD variant",
        tests_realizados="historical local GPU variant of sample-aware MMD",
        best_result="Negative family result",
        latest_result="Closed with the rest of the MMD line.",
        extent_code="E3",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="Host and runtime matter, but they cannot rescue a fundamentally weak modeling direction by themselves.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_cvae/docs/active/ROUTES_AND_RESULTS.md",
        notes="Included for completeness because students will see multiple branch variants of the same strategy family.",
    ),
    BranchLine(
        branch_name="feat/seq-bigru-residual-spline-flow",
        metadata_ref="feat/seq-bigru-residual-spline-flow",
        worktree_path="/workspace/2026/feat_seq_bigru_residual_flow_route",
        strategic_role="Dedicated coupling-flow lane after the historical per-axis flow failure.",
        strategic_hypothesis="A 2D coupling flow could learn a richer residual distribution than the old narrow flow line.",
        architecture_core="seq_bigru_residual",
        generator_family="coupling_2d flow",
        tests_realizados="seq_flow_coupling_smoke; seq_flow_coupling_guided_quick",
        best_result="0/12",
        latest_result="0/12 after guided quick; all four grid candidates were also 0/12 in mini-protocol.",
        extent_code="E4",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="A more expressive density model must still match the project’s residual structure; otherwise large grids only confirm the mismatch.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_flow_route/PROJECT_STATUS.md",
        notes="No undertraining and no collapse; this was a genuine negative scientific verdict.",
    ),
    BranchLine(
        branch_name="feat/seq-bigru-residual-spline-flow-v2",
        metadata_ref="feat/seq-bigru-residual-spline-flow-v2",
        worktree_path="/workspace/2026/feat_seq_bigru_residual_spline_flow_v2",
        strategic_role="Second flow lane using spline-based transforms after coupling-flow failed.",
        strategic_hypothesis="Spline flow could capture global residual shape better than coupling flow or the old narrow flow.",
        architecture_core="seq_bigru_residual",
        generator_family="spline_2d flow",
        tests_realizados="seq_flow_spline_smoke; seq_flow_spline_guided_quick",
        best_result="0/12",
        latest_result="0/12; quick grid also 0/12 in mini-protocol.",
        extent_code="E4",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="Changing the flow parameterization alone did not solve the global mismatch; the whole family is effectively closed for this cycle.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_spline_flow_v2/PROJECT_STATUS.md",
        notes="Marked negative formally; next global-family move became diffusion.",
    ),
    BranchLine(
        branch_name="feat/seq-bigru-residual-diffusion",
        metadata_ref="feat/seq-bigru-residual-diffusion",
        worktree_path="/workspace/2026/feat_seq_bigru_residual_diffusion",
        strategic_role="First diffusion route embedded inside the cVAE scaffold.",
        strategic_hypothesis="A global generative family change to diffusion could solve the residual-shape gap that MDN/flow did not solve.",
        architecture_core="seq_bigru_residual",
        generator_family="diffusion v1: cVAE + diffusion + KL",
        tests_realizados="seq_diffusion_smoke; seq_diffusion_guided_quick",
        best_result="0/12",
        latest_result="0/12 after guided quick; collapse mitigated, but the family remained negative.",
        extent_code="E3",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="It is possible to integrate diffusion structurally and still fail scientifically; collapse is not the only failure mode.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_diffusion/PROJECT_STATUS.md",
        notes="Important teaching branch for separating plumbing success from scientific success.",
    ),
    BranchLine(
        branch_name="feat/seq-bigru-residual-diffusion-v2",
        metadata_ref="feat/seq-bigru-residual-diffusion-v2",
        worktree_path="/workspace/2026/feat_seq_bigru_residual_diffusion_v2",
        strategic_role="Direct conditional diffusion route that removes or sharply weakens the old KL path.",
        strategic_hypothesis="Direct conditional residual diffusion with v-pred/x0-pred could address the global noise-shape mismatch better than diffusion v1.",
        architecture_core="seq_bigru_residual",
        generator_family="diffusion v2: direct conditional residual diffusion",
        tests_realizados="implementation tests (112 passed); seq_diffusion_v2_large with 16 candidates spanning v-pred and x0-pred",
        best_result="1/12 on exp_20260330_114643",
        latest_result="1/12; best v-pred beat best x0-pred, but the family is still scientifically negative.",
        extent_code="E4",
        verdict_flag="NEGATIVE_CLOSED",
        digital_twin_lesson="A major formulation change can remove collapse and undertraining while still proving that the family does not model the digital twin correctly.",
        docs_entrypoint="/workspace/2026/feat_seq_bigru_residual_diffusion_v2/PROJECT_STATUS.md",
        notes="All 16 candidates had active_dim_ratio=1.0 and flag_unstable=True; the launch log was not persisted due to a missing _launch_logs directory at tee startup.",
    ),
]


ARCHITECTURE_LINES = [
    ArchitectureLine(
        architecture_family="concat / channel_residual / delta_residual",
        name_breakdown_and_data_use="Nomes-base da família ponto-a-ponto. `concat`: concatena as entradas e condições diretamente no decoder, sem encoder temporal. `channel_residual`: aprende uma correção residual sobre o canal/base. `delta_residual`: aprende diretamente o delta/resíduo do alvo. Usa amostras independentes por passo, com sinal/contexto local e variáveis de regime (`d`, `c`), sem janela temporal.",
        main_branches="core repo codebase; no dedicated current branch",
        objective_for_digital_twin="Point-wise baselines and supporting families for comparison, ablation, and protocol sanity.",
        interventions_tested="Baseline point-wise cVAE, residual decoder, explicit delta-residual training",
        farthest_stage_code="E5",
        best_verified_result="Used as core baselines/support; not the main open scientific gap in the current cycle.",
        current_status="SUPPORT_BASELINE",
        what_we_learned="Good teaching families for explaining the project stack, but the hard digital-twin gap now lives in the sequence-aware families.",
        reopen_rule="Use as baseline/support, not as the main new scientific route unless a new hypothesis is explicit.",
    ),
    ArchitectureLine(
        architecture_family="seq_bigru_residual + Gaussian",
        name_breakdown_and_data_use="Exemplo representativo: `seq_bigru_residual` com saída Gaussiana. `seq`: usa sequência/janela temporal `W`. `bigru`: GRU bidirecional lê contexto temporal dentro da janela. `residual`: modela o resíduo/ruído condicional, não só o valor absoluto. `Gaussian`: o decoder prevê média e escala de uma distribuição Gaussiana. Usa janelas temporais contíguas mais regime (`d`, `c`) para prever uma distribuição simples por passo.",
        main_branches="historical main line; anchors documented in mdn/diffusion branches",
        objective_for_digital_twin="Stable temporal baseline with simple output family.",
        interventions_tested="Temporal conditioning, protocol-first evaluation, full 12-regime validation",
        farthest_stage_code="E5",
        best_verified_result="10/12 (exp_20260324_023558)",
        current_status="STABLE_REFERENCE",
        what_we_learned="A simple family can still be a strong benchmark; every richer family must beat this line honestly.",
        reopen_rule="Keep as reference and sanity check.",
    ),
    ArchitectureLine(
        architecture_family="seq_bigru_residual + MDN",
        name_breakdown_and_data_use="Exemplo de branch: `feat/seq-bigru-residual-mdn-route`. `seq`: janela temporal. `bigru`: encoder temporal bidirecional. `residual`: aprende o ruído/resíduo condicional ao contexto. `MDN`: Mixture Density Network, mistura de Gaussianas para capturar multimodalidade, caudas e assimetria. `route`: branch dedicada a uma rota experimental. Usa janelas temporais do sinal + condições de regime (`d`, `c`) para aprender `p(resíduo | contexto)`.",
        main_branches="feat/mdn-g5-recovery, feat/seq-bigru-residual-mdn-route",
        objective_for_digital_twin="Current strongest learned family for modeling conditional residual shape.",
        interventions_tested="coverage, tail levels, kurtosis, mdn_components, cond_embed, tuned cond_embed, throughput/runtime variants",
        farthest_stage_code="E5",
        best_verified_result="10/12 (exp_20260328_153611, S27)",
        current_status="ANCHOR_OPEN",
        what_we_learned="This is the best current digital twin line, but it still leaves a systematic global shape mismatch even when protocol passes are high.",
        reopen_rule="Yes, but only with explicit hypotheses tied to the remaining global shape gap.",
    ),
    ArchitectureLine(
        architecture_family="seq_imdd_graybox + Gaussian",
        name_breakdown_and_data_use="Exemplo representativo: `seq_imdd_graybox` Gaussiano. `seq`: janela temporal. `IMDD`: hipótese/estrutura do canal IMDD. `graybox`: mistura de bloco físico conhecido com bloco aprendido. `Gaussian`: saída residual Gaussiana. Usa janelas temporais e variáveis físicas/regime para somar um termo aprendido a um esqueleto físico do canal.",
        main_branches="feat/imdd-graybox-channel",
        objective_for_digital_twin="Physics-guided temporal family using an IMDD gray-box scaffold.",
        interventions_tested="capacity sweeps, guided large sweeps, protocol-first comparison",
        farthest_stage_code="E4",
        best_verified_result="6/12 (exp_20260327_172148)",
        current_status="VIABLE_BELOW_ANCHOR",
        what_we_learned="Physics scaffolding helps but did not reach the MDN seq anchor.",
        reopen_rule="Only with a new clear structural reason.",
    ),
    ArchitectureLine(
        architecture_family="seq_imdd_graybox + MDN",
        name_breakdown_and_data_use="Exemplo de branch: `feat/seq-imdd-graybox-mdn`. `seq`: janela temporal. `IMDD`: estrutura física do canal IMDD. `graybox`: parte física + parte aprendida. `MDN`: mistura de Gaussianas na saída residual. Usa janelas temporais e condições físicas/regime para prever uma distribuição residual mais rica sobre o scaffold gray-box.",
        main_branches="feat/seq-imdd-graybox-mdn",
        objective_for_digital_twin="Test whether MDN on gray-box route fixes the gray-box Gaussian limitation.",
        interventions_tested="smoke, guided quick, full protocol",
        farthest_stage_code="E4",
        best_verified_result="5/12 (exp_20260328_023302)",
        current_status="IMPLEMENTED_NOT_PROMOTED",
        what_we_learned="Implementation is valid, but the route still underperforms both the seq MDN anchor and the best gray-box Gaussian point.",
        reopen_rule="Not a priority unless a new gray-box hypothesis appears.",
    ),
    ArchitectureLine(
        architecture_family="conditional flow decoder (historical per-axis)",
        name_breakdown_and_data_use="Exemplo de branch: `feat/conditional-flow-decoder`. `conditional`: a distribuição depende do contexto de entrada/regime. `flow`: usa transformações invertíveis para modelar densidade. `decoder`: troca a cabeça de saída do modelo. Nesta versão histórica, o flow era estreito e por eixo (`I` e `Q` tratados de forma pouco acoplada). Usava o mesmo contexto residual/temporal das rotas base, mas com uma família generativa mais rígida do que parecia no nome.",
        main_branches="feat/conditional-flow-decoder",
        objective_for_digital_twin="Introduce richer conditional density modeling beyond Gaussian/MDN.",
        interventions_tested="historical protocol experiments with per-axis flow",
        farthest_stage_code="E4",
        best_verified_result="0/12",
        current_status="NEGATIVE_CLOSED",
        what_we_learned="Per-axis narrow flow was too weak for the real residual geometry.",
        reopen_rule="Do not reopen this exact family.",
    ),
    ArchitectureLine(
        architecture_family="seq_bigru_residual + flow (coupling_2d)",
        name_breakdown_and_data_use="Exemplo representativo: rota `flow` com base `seq_bigru_residual`. `seq/bigru/residual`: mesma leitura temporal do sinal e do resíduo. `flow`: distribuição modelada por transformações invertíveis. `coupling_2d`: coupling layers tentando modelar `I/Q` juntos em 2D. Usa janelas temporais + regime como condição e transforma ruído base em residual previsto via flow condicionado.",
        main_branches="feat/seq-bigru-residual-spline-flow",
        objective_for_digital_twin="More expressive 2D density model for residual generation.",
        interventions_tested="structural smoke plus guided quick grid",
        farthest_stage_code="E4",
        best_verified_result="0/12",
        current_status="NEGATIVE_CLOSED",
        what_we_learned="The family was tested seriously enough to rule out a simple tuning explanation.",
        reopen_rule="Only if the whole flow formulation changes again.",
    ),
    ArchitectureLine(
        architecture_family="seq_bigru_residual + flow (spline_2d)",
        name_breakdown_and_data_use="Exemplo representativo: `spline_2d` sobre a base `seq_bigru_residual`. `seq`: janela temporal. `bigru`: encoder temporal bidirecional. `residual`: alvo é o erro/ruído residual. `flow`: densidade gerativa invertível. `spline_2d`: transformações spline em 2D para dar mais flexibilidade ao shape de `I/Q`. Usa janelas + regime e tenta aprender a distribuição residual por um flow spline condicionado.",
        main_branches="feat/seq-bigru-residual-spline-flow-v2",
        objective_for_digital_twin="Second 2D flow attempt using spline transforms.",
        interventions_tested="smoke plus guided quick grid",
        farthest_stage_code="E4",
        best_verified_result="0/12",
        current_status="NEGATIVE_CLOSED",
        what_we_learned="Switching to spline transforms alone did not rescue the flow family.",
        reopen_rule="Closed for this cycle.",
    ),
    ArchitectureLine(
        architecture_family="seq_bigru_residual + diffusion v1",
        name_breakdown_and_data_use="Exemplo de branch: `feat/seq-bigru-residual-diffusion`. `seq`: janela temporal. `bigru`: encoder temporal bidirecional. `residual`: alvo é o resíduo/ruído. `diffusion`: geração por processo de ruído reverso (denoising). `v1`: primeira formulação, ainda acoplada ao cVAE com trilha KL. Usa janelas temporais + regime para condicionar a remoção progressiva de ruído até chegar ao residual previsto.",
        main_branches="feat/seq-bigru-residual-diffusion",
        objective_for_digital_twin="Change the generative family globally via diffusion while keeping the cVAE scaffold.",
        interventions_tested="smoke and guided quick with KL path intact",
        farthest_stage_code="E3",
        best_verified_result="0/12",
        current_status="NEGATIVE_CLOSED",
        what_we_learned="The route integrated technically, but the cVAE+KL scaffold still left the wrong residual family.",
        reopen_rule="Do not retune; treat as closed.",
    ),
    ArchitectureLine(
        architecture_family="seq_bigru_residual + diffusion v2",
        name_breakdown_and_data_use="Exemplo de branch: `feat/seq-bigru-residual-diffusion-v2`. `seq/bigru/residual`: mesma base temporal residual. `diffusion`: geração condicional via denoising. `v2`: segunda formulação, direta, reduzindo/removendo a dependência da trilha latente com KL. Usa janelas temporais + regime para prever o residual por `v-pred` ou `x0-pred`, tentando modelar a distribuição global do ruído de forma mais direta.",
        main_branches="feat/seq-bigru-residual-diffusion-v2",
        objective_for_digital_twin="Direct conditional residual diffusion without depending on the old latent-KL path.",
        interventions_tested="direct implementation, v-pred/x0-pred support, 16-candidate large grid",
        farthest_stage_code="E4",
        best_verified_result="1/12 (exp_20260330_114643)",
        current_status="NEGATIVE_CLOSED",
        what_we_learned="Removing collapse and undertraining still did not make the family scientifically correct for the digital twin.",
        reopen_rule="Only with a new formulation, not with a local retune.",
    ),
    ArchitectureLine(
        architecture_family="sample-aware MMD / regime weighting",
        name_breakdown_and_data_use="Exemplos: `sample-aware-mmd`, `regime weighting`. `sample-aware`: a perda olha para amostras individuais com pesos explícitos. `MMD`: Maximum Mean Discrepancy, métrica/perda de distribuição. `regime weighting`: repondera regimes mais difíceis. Aqui o principal não é trocar a arquitetura de saída, e sim mudar como os dados são amostrados e como a perda força alinhamento de distribuição entre real e sintético.",
        main_branches="feat/sample-aware-mmd*, historical regime-resampling branches",
        objective_for_digital_twin="Attack the residual shape mismatch through loss weighting or resampling rather than changing the generator family.",
        interventions_tested="sample-aware MMD variants; pure regime weighting / resampling",
        farthest_stage_code="E3",
        best_verified_result="Negative family verdict",
        current_status="NEGATIVE_CLOSED",
        what_we_learned="Loss-side pressure alone did not solve the hard G5/global-shape gap.",
        reopen_rule="Do not reopen as the main route.",
    ),
    ArchitectureLine(
        architecture_family="decoder cond_embed within seq MDN",
        name_breakdown_and_data_use="Exemplo: S30/S31 na `feat/mdn-g5-recovery`. `decoder`: mudança localizada na cabeça de saída. `cond_embed`: embedding aprendido da condição/regime entregue ao decoder. `within seq MDN`: preserva o backbone `seq_bigru_residual + MDN` e só especializa a forma como o regime entra no decoder. Usa as mesmas janelas temporais + regime, mas transforma o regime em vetor aprendido para influenciar diretamente a distribuição residual prevista.",
        main_branches="origin/feat/mdn-g5-recovery, feat/mdn-g5-recovery-run",
        objective_for_digital_twin="Decoder-side regime conditioning to attack the remaining low-current G5 shape failures.",
        interventions_tested="S30 cond_embed sweep, S31 tuned cond_embed, S32 large embed sweep",
        farthest_stage_code="E4",
        best_verified_result="Strong δJB improvement but no new full-protocol winner yet",
        current_status="OPEN_SIGNAL_NOT_CLOSED",
        what_we_learned="This is the strongest architecture-side positive signal after S27, but it has not yet crossed the benchmark ceiling.",
        reopen_rule="Yes, but carefully and on the main mdn-g5-recovery line only.",
    ),
]


EXPERIMENT_ROWS = [
    ExperimentRow("exp_20260324_023558", "Stable Gaussian reference", "seq_bigru_residual + Gaussian", "reference", "Stable temporal reference for protocol comparison", "10/12", "G3=10, G5=11, G6=10", "ANCHOR", "Use this as the benchmark every richer family must beat."),
    ExperimentRow("exp_20260325_230938", "Historical MDN line", "seq_bigru_residual + MDN", "historical best", "Historical strong MDN anchor", "9/12", "G3=12, G5=9, G6=12", "ANCHOR", "Important for teaching that MDN can be very strong when the stack is right."),
    ExperimentRow("exp_20260327_161311", "Previous-branch MDN benchmark", "seq_bigru_residual + MDN", "overnight validated rerun", "Best validated prior-branch MDN benchmark", "9/12", "G3=11, G5=10, G6=12", "ANCHOR", "Main reproducibility target for later reruns."),
    ExperimentRow("exp_20260327_172148", "Gray-box Gaussian", "seq_imdd_graybox + Gaussian", "capacity quick", "Best gray-box Gaussian anchor", "6/12", "G3=9, G5=11, G6=6", "VIABLE", "Physics-guided but clearly below the strongest seq MDN anchor."),
    ExperimentRow("exp_20260327_183153", "Gray-box Gaussian", "seq_imdd_graybox + Gaussian", "guided large", "Open gray-box context/capacity more widely", "5/12", "G3=10, G5=9, G6=6", "BELOW_ANCHOR", "More budget removed undertraining but not the protocol gap."),
    ExperimentRow("exp_20260328_003030", "Gray-box + MDN", "seq_imdd_graybox + MDN", "smoke", "Structural integration check", "0/12", "smoke only", "STRUCTURAL_ONLY", "Not a scientific verdict; proves the branch works end-to-end."),
    ExperimentRow("exp_20260328_023302", "Gray-box + MDN", "seq_imdd_graybox + MDN", "guided quick", "First scientific check of gray-box + MDN", "5/12", "G3=9, G5=9, G6=5", "IMPLEMENTED_NOT_PROMOTED", "Valid route, but not better than gray-box Gaussian."),
    ExperimentRow("exp_20260328_041729", "Dedicated MDN rerun lane", "seq_bigru_residual + MDN", "overnight rerun", "Reproduce the historical MDN line under current stack", "4/12", "G3=9, G5=6, G6=5", "REPRO_GAP", "The old anchor did not reproduce cleanly."),
    ExperimentRow("exp_20260328_181213", "Single-candidate MDN isolation", "seq_bigru_residual + MDN", "isolation rerun", "Check one exact historical candidate in isolation", "2/12", "G3=4, G5=8, G6=2", "REPRO_GAP", "Confirms the runtime/environment issue is real."),
    ExperimentRow("exp_20260328_153611", "Main MDN anchor", "seq_bigru_residual + MDN", "S27 lambda_coverage sweep", "Best current seq MDN line", "10/12", "G5=10, G6=12", "ANCHOR", "Current best learned digital twin line."),
    ExperimentRow("exp_20260328_191811", "MDN G5 attack", "seq_bigru_residual + MDN", "S28 lambda_kurt sweep", "Test kurtosis-specific pressure on G5", "negative", "mini regression", "NEGATIVE", "Did not solve G5; could even become catastrophic."),
    ExperimentRow("exp_20260328_210953", "MDN G5 attack", "seq_bigru_residual + MDN", "S29 mdn_components sweep", "Increase MDN capacity globally", "mixed / no fix", "mini 5-8/12 depending config", "NEGATIVE", "More components did not fix the hard 0.8m low-current gap."),
    ExperimentRow("exp_20260328_234446", "MDN G5 attack", "seq_bigru_residual + MDN + cond_embed", "S30 cond_embed sweep", "Test small shared decoder embedding by regime", "4/12 mini best", "δJB improved sharply", "POSITIVE_SIGNAL", "Strong local shape signal, but not a new benchmark."),
    ExperimentRow("exp_20260329_222127", "MDN G5 attack", "seq_bigru_residual + MDN + cond_embed", "S31 tuned cond_embed", "Stability and LR follow-up for cond_embed", "0/12 full", "stable e64/lr2e4 but floor persists", "NEGATIVE_WITH_SIGNAL", "Confirmed the direction is interesting, but the family still hit the same floor."),
    ExperimentRow("exp_20260326_034522", "Historical flow", "conditional flow decoder (sinh-arcsinh)", "historical run A", "Test narrow conditional flow decoder", "0/12", "closed negative", "NEGATIVE", "Historical negative baseline for richer decoder families."),
    ExperimentRow("exp_20260326_035723", "Historical flow", "conditional flow decoder (sinh-arcsinh)", "historical run B", "Repeat narrow conditional flow decoder", "0/12", "closed negative", "NEGATIVE", "Confirms the family verdict."),
    ExperimentRow("exp_20260328_204607", "Flow coupling", "seq_bigru_residual + flow (coupling_2d)", "smoke", "Structural check for coupling flow route", "0/12", "smoke only", "STRUCTURAL_ONLY", "Pipeline plumbing worked."),
    ExperimentRow("exp_20260328_210003", "Flow coupling", "seq_bigru_residual + flow (coupling_2d)", "guided quick", "First scientific grid for coupling flow", "0/12", "mini also 0/12", "NEGATIVE", "Strong enough to close the family for now."),
    ExperimentRow("exp_20260329_015508", "Flow spline", "seq_bigru_residual + flow (spline_2d)", "smoke", "Structural check for spline flow route", "0/12", "smoke only", "STRUCTURAL_ONLY", "Pipeline plumbing worked."),
    ExperimentRow("exp_20260329_015815", "Flow spline", "seq_bigru_residual + flow (spline_2d)", "guided quick", "First scientific grid for spline flow", "0/12", "mini also 0/12", "NEGATIVE", "Enough evidence to close the family for this cycle."),
    ExperimentRow("exp_20260329_210444", "Diffusion v1", "seq_bigru_residual + diffusion v1", "smoke", "Structural check for diffusion v1", "0/12", "collapse signal", "STRUCTURAL_ONLY", "Important because it separates plumbing success from scientific success."),
    ExperimentRow("exp_20260329_211418", "Diffusion v1", "seq_bigru_residual + diffusion v1", "guided quick", "Mitigate collapse and retest diffusion v1", "0/12", "active_dim_ratio=1.0 but still wrong", "NEGATIVE", "Closed the cVAE+diffusion+KL line."),
    ExperimentRow("exp_20260330_114643", "Diffusion v2", "seq_bigru_residual + diffusion v2", "seq_diffusion_v2_large", "Test direct conditional diffusion with v/x0 targets in a 16-candidate grid", "1/12", "G5=3, G6=2", "NEGATIVE", "Better than v1 only marginally; still negative as a digital twin family."),
]


FLAGS_GUIDE = [
    ("E0", "Idea / note only", "No serious implementation or experiment yet."),
    ("E1", "Branch opened / docs only", "Strategy exists conceptually, but has not reached real protocol testing."),
    ("E2", "Structural implementation + smoke", "Code path exists and ran a smoke, but no meaningful scientific comparison yet."),
    ("E3", "Guided quick exploration", "Small scientific grid or focused test exists; enough to see if the idea has signal."),
    ("E4", "Large sweep / dedicated rerun / serious verdict", "The family has been explored far enough to support a real keep/kill decision."),
    ("E5", "Benchmark anchor / mature line", "This line is a strong reference or the current main scientific anchor."),
    ("ANCHOR", "Best current reference", "Keep this visible for students as the line others must beat."),
    ("ANCHOR_OPEN", "Strong anchor, still scientifically open", "Current best line, but not yet the end of the story."),
    ("IMPLEMENTED_NOT_PROMOTED", "Works, but not chosen as main lane", "Useful branch to study architecture variants without promoting them."),
    ("REPRO_GAP", "Main lesson is reproducibility drift", "Good for teaching environment/host/stack provenance."),
    ("NEGATIVE_CLOSED", "Closed negative family", "Do not reopen without a new formulation change."),
    ("NEGATIVE_WITH_SIGNAL", "Negative overall, but contains a useful local clue", "Good teaching case: not all negatives are worthless."),
    ("SNAPSHOT_SUPERSEDED", "Documentation / handoff snapshot", "Useful historically, but not the newest main branch."),
    ("SUPPORT_BASELINE", "Supporting family", "Important for comparison and pedagogy, not the main current research front."),
    ("ROOT_RESEARCH_LINE", "Parent research lane", "A branch that generated later specialized branches and should be taught as lineage, not only as one experiment."),
    ("ARCHIVED_TRACEABILITY", "Archived for traceability", "Shows an idea family existed, but is not an active current-cycle route."),
]


GUIDE_ROWS = [
    ("Workbook purpose", "Explain branch strategy, architecture coverage, and key experiments for beginner students onboarding into the VLC digital twin project."),
    ("Final objective", "Learn a data-driven digital twin that matches the conditional distribution p(y | x, d, c), not only the mean mapping."),
    ("How to read the workbook", "Start with the Branches sheet, then filter by Verdict_Flag and Extent_Code. Use the Architectures sheet to see how far each family was explored."),
    ("Naming grammar", "`feat/` usually marks a feature/research branch; `seq` means temporal windowed data; `bigru` means bidirectional GRU; `residual` means the model predicts residual/noise structure; `mdn` means Mixture Density Network; `route` means a dedicated experimental lane; `v1/v2` means major formulation revisions."),
    ("Important project lesson", "A branch is not only code: it is a scientific hypothesis, an operational environment, and a protocol verdict."),
    ("What counts as a line of work", "This workbook includes active research branches, historical negative branches, root lineage branches, and platform/release branches that matter pedagogically."),
    ("Host provenance", "Some anchors came from a 5090 + tf25_gpu + tmux lane; other reruns came from the local A6000. Students must compare host + branch + stack together."),
    ("Key beginner check", "Before proposing a new branch, verify whether the family is already ANCHOR, OPEN, REPRO_GAP, or NEGATIVE_CLOSED."),
    ("Recommended first docs", "/workspace/2026/feat_seq_bigru_residual_cvae/docs/active/ROUTES_AND_RESULTS.md ; /workspace/2026/feat_seq_bigru_residual_cvae/docs/reference/CONDITIONAL_DENSITY_DECODER_GUIDE.md ; /workspace/2026/feat_seq_bigru_residual_diffusion_v2/docs/reference/CONDITIONAL_DIFFUSION_GUIDE.md"),
]


def git_meta(ref: str) -> tuple[str, str, str]:
    try:
        out = subprocess.check_output(
            [
                "git",
                "-C",
                str(REPO_ROOT),
                "log",
                "-1",
                "--date=iso-strict",
                "--format=%h|%ad|%s",
                ref,
            ],
            text=True,
        ).strip()
        short, date, subject = out.split("|", 2)
        return short, date, subject
    except Exception:
        return "", "", ""


def autosize(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 45)


TOKEN_MEANINGS = {
    "feat": "feature/research branch",
    "exp": "experimental engineering base branch",
    "release": "release-oriented branch",
    "seq": "uses temporal sequences/windows",
    "bigru": "bidirectional GRU temporal encoder",
    "residual": "models residual/noise correction",
    "mdn": "Mixture Density Network decoder",
    "route": "dedicated experimental lane",
    "run": "snapshot/run-specific lane",
    "graybox": "hybrid physical + learned model",
    "imdd": "IMDD physical/channel hypothesis",
    "channel": "channel-focused architecture",
    "recovery": "attempt to recover a failed metric/benchmark",
    "g5": "targets Gate G5, the hard distribution-shape gate",
    "delta": "predicts delta/change or delta-residual target",
    "adv": "adversarial idea/family",
    "cvae": "conditional variational autoencoder core",
    "online": "functional online/serving path",
    "refactor": "engineering refactor / pipeline cleanup",
    "architecture": "architecture-level organization of the code/model",
    "conditional": "conditioned on context/regime",
    "flow": "flow-based density model family",
    "decoder": "change centered on the decoder/output family",
    "sample": "sample-level weighting or sampling logic",
    "aware": "objective explicitly reacts to sample-wise differences",
    "mmd": "Maximum Mean Discrepancy distribution-matching objective",
    "gpu": "GPU-oriented runtime variant",
    "local": "local-machine runtime variant",
    "spline": "spline-based transform family",
    "diffusion": "diffusion/denoising generative family",
    "v2": "second major formulation/revision",
}


def explain_branch_name(branch_name: str) -> str:
    tokens = [tok for tok in re.split(r"[/_-]+", branch_name) if tok]
    parts = []
    for token in tokens:
        meaning = TOKEN_MEANINGS.get(token.lower())
        if meaning:
            parts.append(f"`{token}`: {meaning}")
    prefix = ""
    if branch_name.startswith("feat/"):
        prefix = "This is a research/feature branch name: "
    elif branch_name.startswith("exp/"):
        prefix = "This is an engineering experiment/base branch name: "
    elif branch_name.startswith("release/"):
        prefix = "This is a release/stabilization branch name: "
    if not parts:
        return "Project-specific branch label; read the strategy columns for its role."
    return prefix + "; ".join(parts) + "."


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def append_rows(ws, headers: Iterable[str], rows: Iterable[Iterable[object]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    style_sheet(ws)


def build_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_guide = wb.create_sheet("Guide")
    append_rows(
        ws_guide,
        ["Topic", "Explanation"],
        GUIDE_ROWS,
    )

    ws_flags = wb.create_sheet("Flags")
    append_rows(
        ws_flags,
        ["Flag", "Meaning", "How students should read it"],
        FLAGS_GUIDE,
    )

    branch_rows = []
    for row in BRANCH_LINES:
        short, date, subject = git_meta(row.metadata_ref)
        branch_rows.append(
            [
                row.branch_name,
                explain_branch_name(row.branch_name),
                row.metadata_ref,
                row.worktree_path,
                short,
                date,
                subject,
                row.strategic_role,
                row.strategic_hypothesis,
                row.architecture_core,
                row.generator_family,
                row.tests_realizados,
                row.best_result,
                row.latest_result,
                row.extent_code,
                row.verdict_flag,
                row.digital_twin_lesson,
                row.docs_entrypoint,
                row.notes,
            ]
        )
    ws_branches = wb.create_sheet("Branches")
    append_rows(
        ws_branches,
        [
            "Branch",
            "Name_Tokens_And_Branch_Meaning",
            "Metadata_Ref",
            "Worktree_or_Location",
            "Last_Commit",
            "Last_Commit_Date_UTC",
            "Last_Commit_Subject",
            "Strategic_Role",
            "Strategic_Hypothesis",
            "Architecture_Core",
            "Generator_Family",
            "Tests_Realizados",
            "Best_Result",
            "Latest_Result",
            "Extent_Code",
            "Verdict_Flag",
            "Digital_Twin_Lesson",
            "Docs_Entrypoint",
            "Notes",
        ],
        branch_rows,
    )

    ws_arch = wb.create_sheet("Architectures")
    append_rows(
        ws_arch,
        [
            "Architecture_Family",
            "Name_Tokens_And_Data_Use",
            "Main_Branches",
            "Objective_For_Digital_Twin",
            "Interventions_Tested",
            "Farthest_Stage_Code",
            "Best_Verified_Result",
            "Current_Status",
            "What_We_Learned",
            "Reopen_Rule",
        ],
        [
            [
                row.architecture_family,
                row.name_breakdown_and_data_use,
                row.main_branches,
                row.objective_for_digital_twin,
                row.interventions_tested,
                row.farthest_stage_code,
                row.best_verified_result,
                row.current_status,
                row.what_we_learned,
                row.reopen_rule,
            ]
            for row in ARCHITECTURE_LINES
        ],
    )

    ws_exp = wb.create_sheet("Key_Experiments")
    append_rows(
        ws_exp,
        [
            "Run_ID",
            "Branch_Line",
            "Architecture_Family",
            "Preset_Or_Probe",
            "Purpose",
            "Pass_Result",
            "Gates",
            "Verdict",
            "Note",
        ],
        [
            [
                row.run_id,
                row.branch_line,
                row.architecture_family,
                row.preset_or_probe,
                row.purpose,
                row.pass_result,
                row.gates,
                row.verdict,
                row.note,
            ]
            for row in EXPERIMENT_ROWS
        ],
    )

    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
